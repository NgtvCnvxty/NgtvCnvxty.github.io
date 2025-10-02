################################################################################
### IMPORT STATEMENTS
################################################################################
import requests  # for sending get request to URL
from bs4 import BeautifulSoup  # for parsing URL HTML data
from openpyxl import load_workbook  # for read/write to Excel
from datetime import datetime, date, timedelta  # for dates
import sys  # for exit if failed tests
import re


################################################################################
### CONSTANTS
################################################################################

URL = 'https://www.bankofcanada.ca/rates/interest-rates/canadian-bonds/'
EXCEL_FILE = 'GOC.xlsx'  # In same MarketData project folder

################################################################################
### GOC.xlsx
################################################################################

# --- LOAD SPREADSHEET AND TABS ---
wb = load_workbook(EXCEL_FILE, data_only=True)  # for getting cell values not formulas
ws1 = wb['Bonds']
ws2 = wb['Curve']
# x = (ws1['A1'].value) # for getting specific cell element

# ---  GET ALL DATES FROM EXCEL COLUMNS AND CALCULATE MAX AND LAST ROW---
ws1_dates = [cell.value.date() for cell in ws1['A'][1:]
             if isinstance(cell.value, (datetime, date))]  # skip headers
max_date_ws1 = max(ws1_dates)
ws2_dates = [cell.value.date() for cell in ws2['A'][1:]
             if isinstance(cell.value, (datetime, date))]  # skip headers
max_date_ws2 = max(ws2_dates)

# Check same max dates in file
if max_date_ws1 == max_date_ws2:
    print("OK: Max Date in", EXCEL_FILE, "tabs Bonds/Curves the Same =",
          max_date_ws1.strftime("%Y-%m-%d"))
    max_date_xlsx = max_date_ws1
else:
    print("ERROR: Max date in Bonds != Curve:",
          max_date_ws1.strftime("%Y-%m-%d"), max_date_ws2.strftime("%Y-%m-%d"))
    sys.exit(1)

# Last row of each tab that contains data
last_row_ws1 = max(i for i, cell in enumerate(ws1['A'][1:], start=2)
                   if cell.value is not None)

last_row_ws2 = max(i for i, cell in enumerate(ws2['A'][1:], start=2)
                   if cell.value is not None)

################################################################################
### SCRAPING BOC WEB DATA
################################################################################

# --- SCRAPE PAGE ---
# Get request for URL
try:
    response = requests.get(URL)  # <class 'bytes'>
    response.raise_for_status()
    print("OK: Get URL Success")
except requests.exceptions.RequestException as e:
    print(f"Request failed: {e}")
    sys.exit(1)

# Pull data from URL; <class 'bs4.BeautifulSoup'> in bytes
soup = BeautifulSoup(response.content, 'html.parser')

################################################################################
### SELECT BENCHMARK BOND YIELDS --- TABLE
################################################################################
# table <class 'bs4.element.Tag'> is the entire table
table = soup.find("table", {"id": "table_daily_2"})
#url_string = (soup.prettify()) # <class 'string'> of website content
#print(url_string)

# --- PARSE 5-DAY DATA TABLE WITH DATES AND YIELDS ---

# Getting the headers in the table (i.e. Series, 2025-09-19, ... , 2025-09-25)
# >>> type(headers) <class 'list'>
# >>> type(headers[i]) <class 'bs4.element.Tag'>
# >>> len(headers) = 5
# first element would be "Series" so would mess up datetime so ignoring [1:]
headers = table.find("thead").find_all("th")[1:]

# Format the headers to datetime dates
# >>> type(dates) <class 'list'>
# >>> type(dates[i]) <class 'datetime.date'>
# >>> len(dates) = 5
# >>> type(headers[i].text) <class 'str'>
#     .strip() removes leading/trailing whitespace
#     .replace() unicode hyphen being replaced for standard "-" needed for strptime
#     datetime....date() converts to type datetime.date
# >>> type(headers[1].text.strip().replace('\u2011', '-')) <class 'str')
dates = [datetime.strptime(i.text.strip().replace(
    '\u2011', '-'),"%Y-%m-%d").date() for i in headers]
max_date_url = max(dates)

# Check that there is new data to update from URL
if max_date_url > max_date_xlsx:
    print("OK: Data to Update in", EXCEL_FILE, "as Max Date in URL",
          max_date_url.strftime("%Y-%m-%d"), "> Max Date in", EXCEL_FILE,
          max_date_xlsx.strftime("%Y-%m-%d"))
    print("\tDates to Update in", EXCEL_FILE)
    num = 1
    for i in dates:
        if i > max_date_xlsx:
            print(f"\t(", num, ")", i)
            num += 1
else:
    print("OK: No Data to Update as Max Date in URL",
          max_date_url.strftime("%Y-%m-%d"), "= Max Date in", EXCEL_FILE)
    sys.exit(1)

# Extract table rows (i.e. 2-year, 2.46, 2.45, ... RRB ... long-term ... 1.71)
# >>> type(rows) <class 'bs4.element.ResultSet'>
# >>> len(rows) = 9
# >>> type(rows[i]) <class 'bs4.element.Tag'>
# >>> rows[1] is 2-year data ... rows[6] is 30-year nominal ... rows[8] is RRB
# >>> rows[0] and rows[7] are "Government of Canada benchmark bond yields" and "Real return bond"
# >>> len(rows[i]) i = 0 or 7 = 1
# >>> len(rows[i]) otherwise = 6
rows = table.find("tbody").find_all("tr")

# Define bond types corresponding to row indices 1, 2, 3, 4, 5, 6, 8
bond_types = ["2-year", "3-year", "5-year", "7-year", "10-year", "30-year", "30-year RRB"]
row_indices = [1, 2, 3, 4, 5, 6, 8]

# Create yield_data dictionary
yield_data = {}
for i, j in zip(row_indices, bond_types):
    # Extract cells from the row, starting from index 1 to match dates
    cells = rows[i].find_all("td")[:]
    # Convert cell text to float for yield values (handle empty strings as None or 0.0 if needed)
    yields = []
    for cell in cells:
        text = cell.text.strip()
        yields.append(float(text) if text else
                      None)  # Use None for empty cells; adjust as needed
    # Create dictionary entry with bond as key and yields list as value
    yield_data[j] = yields
# yield_data has keys as in bond_types
# yield_data["2-year"] = [2.46, 2.45, 2.45, 2.45, 2.49]
# yield_data["2-year"][4] = 2.49
# yield_data.items() = dict_items([('2-year', [2.46, 2.45, 2.45, 2.45, 2.49]), ... , ('30-year RRB', [1.68, 1.69, 1.7, 1.71, 1.71])])
# len(yield_data) = 7
# len(yield_data[key]) = 5

# Print all keys:values in yield_data dictionary
print("OK: All Yield Data at URL")
for bond, yields in yield_data.items():
    print(f"\t{bond}: {yields}")

# --- FILTERING SCRAPED DATA FOR EXCEL_FILE UPDATE ---

# Filtering yield dictionary to remove values for dates < max_date_xlsx
index_max_date = dates.index(max_date_xlsx) + 1  # Indices for which will keep data in dictionary for update
filtered_yield_data = {key: value[index_max_date:] for key, value in yield_data.items()}
print("OK: Filtered Yield Data at URL to Update in", EXCEL_FILE)
for bond, yields in filtered_yield_data.items():
    print(f"\t{bond}: {yields}")

################################################################################
### CURRENT BENCHMARK BOND YIELDS --- LIST
################################################################################

# --- PARSE CURRENT BENCHMARK BOND YIELDS (BOTTOM SECTION OF PAGE) ---

# Extract all paragraph or list elements that may contain bond benchmark data
# We're looking for the section at the bottom of the page with 7 lines like:
# "2 year - 2027.08.01, 2.50% (2025.07.24);"
raw_text = soup.get_text()

# Regex to find lines like: "2 year - 2027.08.01, 2.50%"
pattern = re.compile(r'(?P<term>2 year|3 year|5 year|7 year|10 year|Long|RRB)\s*-\s*(?P<maturity>\d{4}\.\d{2}\.\d{2}),\s*(?P<coupon>\d+\.\d+)%')

# Bond type mapping to match your desired dictionary keys
bond_key_map = {"2 year": "2-year", "3 year": "3-year", "5 year": "5-year", "7 year": "7-year",
                "10 year": "10-year", "Long": "30-year", "RRB": "30-year RRB"}

# Initialize the dictionary
current_benchmarks: dict[str, list] = {}

# Iterate through matches in the raw page text
for match in pattern.finditer(raw_text):
    term, maturity_str, coupon_str = match.groups()
    key = bond_key_map[term]
    maturity = datetime.strptime(maturity_str, "%Y.%m.%d").date()
    coupon = float(coupon_str)
    current_benchmarks[key] = [coupon, maturity]

# Optional: Verify dictionary structure (matches your original request)
print("OK: Current Benchmark Bonds (coupon, maturity):")
for bond, data in current_benchmarks.items():
    print(f"\t{bond}: [{data[0]}%, {data[1].strftime("%Y-%m-%d")}]")

################################################################################
### UPDATING EXCEL_FILE
################################################################################

# --- BOTH TABS ---
dates_to_update = [d for d in dates if d > max_date_xlsx]

# --- CURVE TAB ---
# Filtering dates for those in the URL that will be updated in EXCEL_FILE (protects for weekends)
bond_order = ["2-year", "3-year", "5-year", "7-year", "10-year", "30-year", "30-year RRB"]
curve_update = []
for i, date in enumerate(dates_to_update):
    row = [date]
    for bond in bond_order:
        row.append(filtered_yield_data[bond][i])
    curve_update.append(row)
print("UPDATE: Data in CURVE")
for row in curve_update:
    formatted_date = row[0].strftime("%Y-%m-%d")
    formatted_row = [formatted_date] + row[1:]
    print("\t[" + ", ".join([formatted_date] + [str(x) for x in row[1:]]) + "]")

# Writing to EXCEL_FILE
start_row_curve = last_row_ws2 + 1

for i, row in enumerate(curve_update):
    excel_row = start_row_curve + i
    # Write date to column A
    ws2.cell(row=excel_row, column=1, value=row[0])
    # Write yields to columns B onwards
    for j, val in enumerate(row[1:], start=2):
        ws2.cell(row=excel_row, column=j, value=val)

# --- BONDS TAB ---

# Writing to EXCEL_FILE
start_row_bonds = last_row_ws1 + 1

# Define tenor formatting helper to convert full bond name to tenor label
def extract_tenor(bond_name: str) -> str:
    if "RRB" in bond_name:
        return "RRB"
    return bond_name.split("-")[0]  # e.g., "2-year" → "2"

# Build rows to write to Bonds tab
bonds_update = []
for i, date in enumerate(dates_to_update):
    for bond in bond_order:
        tenor_raw = extract_tenor(bond)
        tenor = int(tenor_raw) if tenor_raw.isdigit() else tenor_raw  # numeric tenor as int
        coupon = current_benchmarks[bond][0]  # float
        maturity = current_benchmarks[bond][1]
        ytm = filtered_yield_data[bond][i]
        bonds_update.append([date, tenor, coupon, maturity, ytm])

# Write to Excel starting at the next empty row
for i, row in enumerate(bonds_update):
    excel_row = start_row_bonds + i
    for j, val in enumerate(row, start=1):  # Columns A to E
        ws1.cell(row=excel_row, column=j, value=val)

# Print bonds update data
print("UPDATE: Data in BONDS")
for row in bonds_update:
    date_str = row[0].strftime("%Y-%m-%d")
    tenor, coupon, maturity, ytm = row[1:]
    maturity_str = maturity.strftime("%Y-%m-%d")
    print(f"\t[{date_str}, {tenor}, {coupon}, {maturity_str}, {ytm}]")

################################################################################
### TEST UPDATES TO EXCEL_FILE
################################################################################

# --- TEST NEW NUMBER OF ROWS OF CURVE AND BONDS CORRECT ---
expected_bonds_rows = last_row_ws1 + len(dates_to_update) * 7
expected_curve_rows = last_row_ws2 + len(dates_to_update)

new_ws1_rows: int = max(i for i, cell in enumerate(ws1['A'][1:], start=2) if cell.value is not None)
new_ws2_rows: int = max(i for i, cell in enumerate(ws2['A'][1:], start=2) if cell.value is not None)

print(f"VERIFY: Curve tab rows → Expected: {expected_curve_rows}, Actual: {new_ws1_rows}")
print(f"VERIFY: Bonds tab rows → Expected: {expected_bonds_rows}, Actual: {new_ws2_rows}")

# --- SAVE FILE ---
wb.save(EXCEL_FILE)
wb.close()
